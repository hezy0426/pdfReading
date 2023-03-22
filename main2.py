import os
import openpyxl
import fnmatch
import re
import cv2
import pytesseract
import numpy
from pdf2image import convert_from_path
from shutil import rmtree
from datetime import date


class ohNo(BaseException):
    PREVIOUSFILE = '最终报价.*'
    key_dict = {
        'name': '短文本',
        'brand': '供应商/供应工厂',
        'note': '备注',
        'productNum': '物料',
        'cost': '含税价',
        'orderAmount': '采购订单数量'
    }
    key_dict2 = {
        'weeklyReport': '月报',
        'MONTHLYORDERTOTAL': '月采购总量（颗）',
        'MONTHLYORDERAMOUNT': '月采购总额（元）',
        'ORDERREQUEST': '采购申请号',
        'MONTHLYTOTALREDUCE': '月降价总额（元）',
        'MONTHLYREDUCEPER': '月降价比例',
        'costSingle': '采购单价（综合）',
        'costPrevious': '去年采购单价',
        'reducedPercent': '降价比例',
        'reducedTotal': '降价总额'
    }

    def __init__(self, outputFolder, inputFolder):
        self.libFolder = os.path.join(os.getcwd(), '资料')
        if not os.path.exists(self.libFolder):
            os.makedirs(self.libFolder)
        print(self.libFolder)

        self.outputPath = outputFolder
        self.inputPath = inputFolder

    # Check if the weeklyreport input file is correctly formatted
    def isPreviousYearFileValid(self, ws1, rowOrder):
        count = 0
        for row in ws1.iter_rows(min_row=1, max_row=1, values_only=True):
            for x in row:
                if x == self.key_dict['productNum']:
                    count += 1
                    rowOrder[self.key_dict['productNum']] = row.index(x)
                if x == self.key_dict['cost']:
                    count += 1
                    rowOrder[self.key_dict['cost']] = row.index(x)

        return not count == 2

    # Create a new dictionary for the a specific year's price. This will search for the specific file name
    # and then grab the value from first column to save as key and third column as value
    def previousYearDataStru(self, yearNum):
        result = ""
        for root, dirs, files in os.walk(self.libFolder):
            for name in files:
                if fnmatch.fnmatch(name, yearNum+self.PREVIOUSFILE):
                    result = (os.path.join(root, name))
        print(result)
        if not result:
            print('找不到对应文件,确认资料文件夹里有这个'+yearNum+'年的文件')
            return None

        price_dict = {}
        rowOrder = {}

        wb_obj = openpyxl.load_workbook(result)

        ws1 = wb_obj.active
        if self.isPreviousYearFileValid(ws1, rowOrder):
            print(yearNum+'最终报价文件第一行有物料或含税价')
            return
        for row in ws1.iter_rows(min_row=2, max_row=ws1.max_row, values_only=True):
            price_dict[row[rowOrder[self.key_dict['productNum']]]
                       ] = row[rowOrder[self.key_dict['cost']]]

        return price_dict

    # Find path for xlsx file in the same folder
    def findXLSX(self):
        result = ""
        for root, dirs, files in os.walk(self.inputPath):
            for name in files:
                if fnmatch.fnmatch(name, '*.xlsx'):
                    result = (os.path.join(root, name))
        print(result)
        if not result:
            print('输入文件夹里没有xlsx文件,请重新确认一遍')
        return result

    # Validate the input string, it must be 4 digits number
    def validateNumber(self, yearNum):
        if re.search("[0-9]{4}", yearNum):
            return True
        return False

    # Try to see if contains "XXB%d%d%d%d", if not, try "%d%d%d%d". If neither is found, just returns 0000
    def getNum1(self, Name):
        temp = re.search("[0-9]{4}", Name)
        if temp is None:
            return "0000"
        return temp.group()

    # Check if the weeklyreport input file is correctly formatted
    def isWeeklyReportNotValid(self, ws1):
        count = 0
        for row in ws1.iter_rows(min_row=1, max_row=1, values_only=True):
            for x in row:
                if x in self.key_dict.values():
                    count += 1

        return not count == len(self.key_dict)

    # Get name and order amount. This will call getNum1() to get the typeNum. Name and order amount will be saved in a dictionary, and everything else (name is also included) will be written directly into the new sheet
    def readRows(self, ws1, ws2, yearNum):
        price_dict = self.previousYearDataStru(yearNum)
        if not price_dict:
            return None

        key_dict = self.key_dict
        allSums = {}

        rowOrder = {}

        for row in ws1.iter_rows(min_row=1, max_row=1, values_only=True):
            for keys in key_dict:
                rowOrder[keys] = row.index(key_dict[keys])

        for row in ws1.iter_rows(min_row=2, max_row=ws1.max_row, values_only=True):
            name = row[rowOrder['name']]
            if name in allSums:
                allSums[name] += row[rowOrder['orderAmount']]
            else:
                allSums[name] = row[rowOrder['orderAmount']]

                tempType = ''

                if '铜线' in name:
                    tempType = '铜线'
                elif '金线' in name:
                    tempType = '金线'

                if row[rowOrder['productNum']] in price_dict:
                    lastYearPrice = price_dict[row[rowOrder['productNum']]]
                    reducedPer = (
                        lastYearPrice-(row[rowOrder['cost']]/1000))/lastYearPrice
                    reducedTotal = 0
                else:
                    lastYearPrice = 'N/A'
                    reducedPer = 'N/A'
                    reducedTotal = 'N/A'

                ws2.append((name, self.getNum1(name), tempType, row[rowOrder['brand']], row[rowOrder['cost']]/1000,
                           lastYearPrice, reducedPer, reducedTotal, '', '', row[rowOrder['note']], '', row[rowOrder['productNum']]))

        return allSums

    # Add final order amount and total cost for that one product
    def calculateAllThreeSums(self, ws2, allSums):
        rowNum = ws2.max_row
        i = 2

        totalNum = 0
        totalCost = 0
        totalReduce = 0

        while (i <= rowNum):
            temp = allSums[ws2['A{}'.format(i)].value]
            ws2['I{}'.format(i)] = temp  # 订单数额
            totalNum += temp

            ws2['J{}'.format(i)] = temp * ws2['E{}'.format(i)].value  # 总额
            totalCost += ws2['J{}'.format(i)].value

            if ws2['H{}'.format(i)].value != 'N/A':
                ws2['H{}'.format(i)] = ws2['F{}'.format(
                    i)].value * temp - ws2['J{}'.format(i)].value
                totalReduce += ws2['H{}'.format(i)].value
            i += 1

        ws2['E{}'.format(rowNum+1)] = self.key_dict2['MONTHLYORDERTOTAL']
        ws2['E{}'.format(rowNum+2)] = self.key_dict2['MONTHLYORDERAMOUNT']
        ws2['E{}'.format(rowNum+3)] = self.key_dict2['MONTHLYTOTALREDUCE']
        ws2['E{}'.format(rowNum+4)] = self.key_dict2['MONTHLYREDUCEPER']
        ws2['F{}'.format(rowNum+1)] = totalNum
        ws2['F{}'.format(rowNum+2)] = totalCost
        ws2['F{}'.format(rowNum+3)] = totalReduce
        ws2['F{}'.format(rowNum+4)
            ] = "{0:.0f}%".format(totalReduce/totalCost * 100)

    # Set up the second sheet, this is where output will go
    def setupOutput(self):
        wb2 = openpyxl.Workbook()
        # Get the first sheet of the new xlsx file
        ws2 = wb2.active
        # Setup a new title
        ws2.title = self.key_dict2['weeklyReport']

        ws2.append((self.key_dict['name'], "灯珠型号", "引线规格", self.key_dict['brand'], self.key_dict2['costSingle'], self.key_dict2['costPrevious'], self.key_dict2['reducedPercent'],
                   self.key_dict2['reducedTotal'], self.key_dict['orderAmount'], "采购总额（元）", self.key_dict['note'], "其他说明", self.key_dict['productNum']))

        ws2.column_dimensions['A'].width = 40
        ws2.column_dimensions['B'].width = 10
        ws2.column_dimensions['D'].width = 36
        ws2.column_dimensions['E'].width = 25
        ws2.column_dimensions['F'].width = 15
        ws2.column_dimensions['G'].width = 15
        ws2.column_dimensions['H'].width = 15
        ws2.column_dimensions['I'].width = 15
        ws2.column_dimensions['J'].width = 20
        ws2.column_dimensions['K'].width = 20
        ws2.column_dimensions['M'].width = 20

        return wb2

    # This is for weekly report.
    def weeklyReport(self, yearNum):
        filePath = self.findXLSX()
        if not filePath:
            return

        # Open the file
        wb_obj = openpyxl.load_workbook(filePath)

        # ws1 is the first page, and ws2 is where we will work on
        ws1 = wb_obj.active

        if self.isWeeklyReportNotValid(ws1):
            print(filePath[(filePath.index('输入\\')+3):] +
                  '的第一行格式不正确,确保第一行包括"短文本","供应商描述","含税价","采购申请号数量",“备注”,"物料"')
            return

        wb2 = self.setupOutput()

        if not ws1 or not wb2.active:
            print("ws1和ws2是null,重试一次把")
            return

        # If we can't find the file in 资料,then it will return none
        allSums = self.readRows(ws1, wb2.active, yearNum)
        if not allSums:  # error msg has been printed up there
            return

        self.calculateAllThreeSums(wb2.active, allSums)
        # Save all the work, since this is a new xlsx file, we need to give it a name
        wb2.save('{0}\\{1}'.format(self.outputPath,
                 date.today().strftime('%Y-%m-%d')+self.key_dict2['weeklyReport']+".xlsx"))

    def setUpLastWeek(self, ws1, lastWeekOrder):
        orderCol = 0
        for i in ws1[1]:
            if i.value == self.key_dict2['ORDERREQUEST']:
                orderCol = i.column_letter

        for i in ws1[orderCol]:
            lastWeekOrder.add(i.value)

        return lastWeekOrder

        # Saving all the processed files
    def savingFILES(self, ws, newOrder):
        for i in newOrder:
            if i == self.key_dict['brand'] or i == None:
                continue
            wb_temp = openpyxl.Workbook()
            ws_temp = wb_temp.active
            ws_temp.append(cell.value for cell in ws[1])
            for x in newOrder[i]:
                ws_temp.append(b.value for b in x)
            wb_temp.save('{0}\\{1}'.format(self.outputPath, str(i)+'.xlsx'))

    def compareToLastWeek(self, wb_obj):
        ws1 = wb_obj.worksheets[0]
        ws2 = wb_obj.worksheets[1]

        # Find the column of the empty letter and order number
        toFindEmptyLetter = ''
        toFindOrder = ''
        companyName = ''
        correctFile = 0

        # Verify if both sheets have the correct columns
        for i in ws1[1]:
            if i.value == self.key_dict2['ORDERREQUEST'] or i.value == '最新回复交期' or i.value == self.key_dict['brand']:
                correctFile += 1
        for i in ws2[1]:
            if i.value == self.key_dict2['ORDERREQUEST']:
                toFindOrder = i.column_letter
            if i.value == '最新回复交期':
                toFindEmptyLetter = i.column_letter
            if i.value == self.key_dict['brand']:
                companyName = i.column_letter

        if not toFindEmptyLetter or not toFindOrder or not companyName or correctFile != 3:
            raise Warning('查看第一行是否有 采购申请号 最新回复交期 供应商/供应工厂')

        # Save all order number from the previous week
        lastWeekOrder = set()
        lastWeekOrder = self.setUpLastWeek(ws1, lastWeekOrder)

        # Save the order number that is found
        newOrder = {}

        # Find the empty String
        for i in ws2[toFindEmptyLetter]:
            if not i.value:  # If this cell is empty
                # If this order number can't be found from last week
                if ws2['{}{}'.format(toFindOrder, i.row)].value not in lastWeekOrder:
                    # If it doesn't exist in the dictionary, create one
                    if ws2['{}{}'.format(companyName, i.row)].value not in newOrder:
                        newOrder[ws2['{}{}'.format(
                            companyName, i.row)].value] = []

                    # Append this dictionary with new value
                    newOrder[ws2['{}{}'.format(companyName, i.row)].value].append(
                        ws2[i.row])

        self.savingFILES(ws2, newOrder)

    def groupAllOrdersFromTheSameProvider(self, wb_obj):
        ws = wb_obj.active

        # Find the column of the provider
        companyName = ''
        for i in ws[1]:
            if i.value == self.key_dict['brand']:
                companyName = i.column_letter
                break

        if not companyName:
            raise Warning('查看第一行是否有供应商/供应工厂')

        # Save the order number that is found
        newOrder = {}

        # Find the empty String
        for i in ws[companyName]:
            # If it doesn't exist in the dictionary, create one
            if ws['{}{}'.format(companyName, i.row)].value not in newOrder:
                newOrder[ws['{}{}'.format(companyName, i.row)].value] = []

            # Append this dictionary with new value
            newOrder[ws['{}{}'.format(companyName, i.row)].value].append(
                ws[i.row])

        self.savingFILES(ws, newOrder)

    # This function will group all orders from the same provider into one xlsx file and save it in the output folder.
    def orderFollowup(self):
        filePath = self.findXLSX()
        if not filePath:
            return

        # Open the file
        wb_obj = openpyxl.load_workbook(filePath)
        if len(wb_obj.worksheets) == 2:
            self.compareToLastWeek(wb_obj)

        elif len(wb_obj.worksheets) == 1:
            self.groupAllOrdersFromTheSameProvider(wb_obj)
        else:
            raise Warning('这个xlsx文件有'+len(wb_obj.worksheets) +
                          '个sheets.目前只支持1个或者2个')

    # This function will read the company name from the PDF file and rename the PDF file with the name
    def nameChange(self):
        # Temp folder for the converted image
        tempHolder = os.path.join(self.outputPath, 'temp')
        if not os.path.exists(tempHolder):
            os.makedirs(tempHolder)

        # First read the pdf file, all paths are saved in a list
        pdfPaths = []
        for root, dirs, files in os.walk(self.inputPath):
            for name in files:
                if fnmatch.fnmatch(name, '*.pdf'):
                    pdfPaths.append(os.path.join(root, name))

        for i in pdfPaths:
            print(i)

        if len(pdfPaths) == 0:
            print('文件夹内找不到PDF文件')
            return

        binLocation = r'poppler-22.01.0\Library\bin'
        exeLocation = r'Tesseract-OCR\tesseract.exe'

        # 1. Read first page| 2. Convert to image| 3. Crop the image to specific size
        # while loop until it's empty
        while len(pdfPaths) > 0:
            tempPath = pdfPaths.pop(0)

            # Needs the library
            pages = convert_from_path(
                tempPath, 500, poppler_path=os.path.join(os.getcwd(), binLocation))

            # Save the first page
            image_name = os.path.join(tempHolder, 'tempImage.jpg')
            pages[0].save(image_name, 'JPEG')

            # fileName = tempPath.split('\\')
            # fileName = fileName[-1].split('.')
            # print(fileName)

            # Crop the saved image to reduce reading time. Now it can even read path with Chinese characters
            try:

                stream = open(image_name, "rb")
                bytes = bytearray(stream.read())
                numpyarray = numpy.asarray(bytes, dtype=numpy.uint8)
                img = cv2.imdecode(numpyarray, cv2.IMREAD_UNCHANGED)
                # img = cv2.imread(image_name)
                roi_color = img[0:1200, 0:4960]
                # roi_color = cv2.cvtColor(roi_color, cv2.COLOR_BGR2GRAY)
                # cv2.imwrite('final{}.jpg'.format(randint(1,20)),roi_color)
            except Exception as err:
                print(err)

            # Read the final image
            try:
                pytesseract.pytesseract.tesseract_cmd = os.path.join(
                    os.getcwd(), exeLocation)  # Location for the OCR
                # Need to specify the language is Chinese. The last param is to remove space between words
                text = str(((pytesseract.image_to_string(
                    roi_color, lang="chi_sim", config='--psm 6 -c preserve_interword_spaces=1'))))
            except Exception as err:
                if 'tesseract is not installed or it\'s not in your PATH.' in err:
                    raise Warning(
                        'Check if the folder has Tesseract-OCR folder')

            # TODO use regex to improve the searching sequence
            print("\n读取的内容:")
            print(text)
            # Parse out the needed info
            orderNum = re.search("订单编号.*[0-9]*", text)
            companyName = re.search("供应商全称.*公司", text)

            # If nothing is found
            if orderNum is None:
                print("找不到订单号")
                continue
            if companyName is None:
                companyName = re.search("供应商全称.*币", text)
                if companyName is None:
                    print('找不到供应商名称')
                    continue

            companyName = re.sub('供应商全称: (.*)', r'\1', companyName.group())

            print(orderNum.group())
            print(companyName)

            # Parse the actual data
            orderNum = re.search("[0-9]+", orderNum.group())
            orderNum = orderNum.group()

            # Remove the first image
            stream.close()
            os.remove(image_name)

            # Once the name is out, rename the file, move the file and remove the path from list
            try:
                os.rename(tempPath, os.path.join(
                    self.outputPath, orderNum+companyName+'.pdf'))
            except Exception as err:
                print(err)

        # Remove the temp folder
        rmtree(tempHolder)

    def testing(self):
        filePath = self.findXLSX()

        # Open the file
        wb_obj = openpyxl.load_workbook(filePath)

        # ws1 is the first page, and ws2 is where we will work on
        ws1 = wb_obj.active
        if self.isFileNotValid(ws1):
            print(filePath[(filePath.index('输入\\')+3):] +
                  '的第一行格式不正确,确保第一行包括"短文本","供应商描述","含税价","采购申请号数量",“备注”,"物料"')
            return
